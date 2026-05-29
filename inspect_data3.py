import openpyxl

wb = openpyxl.load_workbook('VFMC.Daily.20260430.CLIENT_DTF-Act3.AssetClass.xlsx', data_only=True)
ws = wb['RiskReport']

# Row 21 has the actual column headers in the data area
row21 = [cell.value for cell in ws[21]]
print("=== Row 21 headers (all non-empty) ===")
col_map = {}
for i, v in enumerate(row21):
    if v is not None:
        s = str(v).strip()
        if s:
            sl = s.lower()
            if any(k in sl for k in ['level', 'name', 'total', 'asset class', 'manager',
                                      'int_ext', 'iscash', 'sectype', 'portfolio',
                                      'strategy', 'instrument', 'pv aud', 'scd pv',
                                      'security', 'counter', 'trading', 'model',
                                      'client']):
                print(f"  Col {i}: {s}")
            # Map key columns
            if 'level' == sl:
                col_map['level'] = i
            elif sl == 'name':
                col_map['name'] = i
            elif sl == 'total' and 'pv' not in col_map:
                col_map['pv'] = i  # first 'Total' = RM PV AUD
            elif sl == '*vfmc_asset class':
                col_map['ac'] = i
            elif sl == '*vfmc_managername':
                col_map['mgr_name'] = i
            elif sl == '*vfmc_manager':
                col_map['mgr_code'] = i
            elif sl == '*vfmc_int_ext':
                col_map['int_ext'] = i
            elif sl == '*vfmc_sectype':
                col_map['sec_type'] = i
            elif sl == '*vfmc_iscash':
                col_map['is_cash'] = i
            elif sl == 'clientportfolioid':
                col_map['portfolio'] = i
            elif sl == '*vfmc_strategy':
                col_map['strategy'] = i
            elif sl == '*vfmc_asset class category':
                col_map['ac_cat'] = i
            elif sl == '*vfmc_sub asset class':
                col_map['sub_ac'] = i
            elif sl == '*vfmc_clientinvestmentstrategy':
                col_map['cis'] = i
            elif sl == '*vfmc_securityname':
                col_map['sec_name'] = i
            elif sl == '*vfmc_trading_strategy':
                col_map['trading'] = i

print(f"\n=== Column map ===")
for k, v in sorted(col_map.items()):
    print(f"  {k}: col {v} = {row21[v]}")

# Now find AEQ and IEQ data
print(f"\n=== Searching for AEQ/IEQ rows (Level 0) ===")
max_row = ws.max_row
print(f"  Total rows in sheet: {max_row}")

# Scan for Level 0 rows (asset class level)
ac_col = col_map.get('ac')
lv_col = col_map.get('level')
pv_col = col_map.get('pv')
name_col = col_map.get('name')

aeq_rows = []
ieq_rows = []

for row_num in range(22, min(max_row + 1, 200)):
    level = ws.cell(row=row_num, column=lv_col + 1).value
    name = ws.cell(row=row_num, column=name_col + 1).value
    ac = ws.cell(row=row_num, column=ac_col + 1).value if ac_col else None
    pv = ws.cell(row=row_num, column=pv_col + 1).value if pv_col else None
    
    if level == 0:
        name_s = str(name) if name else ''
        print(f"  Row {row_num}: Level={level} Name={name_s} AC={ac} PV={pv}")
        if 'australian equit' in name_s.lower():
            aeq_rows.append(row_num)
        if 'international equit' in name_s.lower() or 'emerging' in name_s.lower() or 'low volatility' in name_s.lower():
            ieq_rows.append(row_num)

# Print some Level 1 children under AEQ or IEQ
if aeq_rows:
    start = aeq_rows[0] + 1
    print(f"\n=== AEQ children (starting row {start}) ===")
    for row_num in range(start, min(start + 30, max_row + 1)):
        level = ws.cell(row=row_num, column=lv_col + 1).value
        if level is not None and level <= 0:
            break
        name = ws.cell(row=row_num, column=name_col + 1).value
        pv = ws.cell(row=row_num, column=pv_col + 1).value
        ac = ws.cell(row=row_num, column=ac_col + 1).value if ac_col else None
        mgr = ws.cell(row=row_num, column=col_map['mgr_name'] + 1).value if 'mgr_name' in col_map else None
        mc = ws.cell(row=row_num, column=col_map['mgr_code'] + 1).value if 'mgr_code' in col_map else None
        ie = ws.cell(row=row_num, column=col_map['int_ext'] + 1).value if 'int_ext' in col_map else None
        st = ws.cell(row=row_num, column=col_map['strategy'] + 1).value if 'strategy' in col_map else None
        
        print(f"  Row {row_num}: Lv={level} Name={name} PV={pv} AC={ac} Mgr={mgr} MC={mc} IE={ie} Strat={st}")

wb.close()
