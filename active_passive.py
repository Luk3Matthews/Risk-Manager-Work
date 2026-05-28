#!/usr/bin/env python3
"""
VFMC Risk Manager — Active/Passive & Internal/External Split Calculator
Reads VFMC RM export .xlsx and classifies AEQ / IEQ managers.
Requires: openpyxl (pip install openpyxl)
"""

import argparse
import csv
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from enum import Enum, auto
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════════════
# Enums
# ═══════════════════════════════════════════════════════════════════════════════

class AssetClassTag(Enum):
    AEQ = auto()
    IEQ = auto()
    OTHER = auto()

class ActivityType(Enum):
    ACTIVE = "Active"
    PASSIVE = "Passive"

class SourceType(Enum):
    INTERNAL = "Internal"
    EXTERNAL = "External"


# ═══════════════════════════════════════════════════════════════════════════════
# Data classes
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class ParsedRow:
    entity_name: str = ""
    asset_class_raw: str = ""
    exposure: float = 0.0
    mandate_code: str = ""
    portfolio: str = ""
    security_type: str = ""


@dataclass
class ManagerEntry:
    name: str = ""
    mandate_code: str = ""
    ac: AssetClassTag = AssetClassTag.OTHER
    act: ActivityType = ActivityType.ACTIVE
    src: SourceType = SourceType.EXTERNAL
    exposure_aud: float = 0.0
    classified: bool = True


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers
# ═══════════════════════════════════════════════════════════════════════════════

def icontains(haystack: str, needle: str) -> bool:
    """Case-insensitive substring check."""
    return needle.lower() in haystack.lower()


def fmt_money(val_m: float) -> str:
    """Format as $X,XXX.Xm"""
    sign = "-" if val_m < 0 else ""
    return f"${sign}{abs(val_m):,.1f}m"


def fmt_pct(pct: float) -> str:
    return f"{pct:.1f}%"


# ═══════════════════════════════════════════════════════════════════════════════
# Column detection (flexible, case-insensitive, partial match)
# ═══════════════════════════════════════════════════════════════════════════════

COLUMN_PATTERNS = {
    "entity": [
        "entity_long_name", "entitylongname", "entity long name",
        "manager_name", "managername", "manager name",
    ],
    "asset_class": [
        "asset_class", "assetclass", "asset class",
        "business_class_level_4", "businessclasslevel4",
    ],
    "exposure": [
        "exposure_pc", "eff_exp_value_base",
        "effectivedateexposure", "effective_date_exposure",
        "exposure", "eff_exp", "mkt_val", "market_value", "marketvalue",
    ],
    "mandate": [
        "scd_mandate", "mandate_code", "mandatecode", "mandate code", "mandate",
    ],
    "portfolio": [
        "portfolio", "trustname", "trust_name", "trust name",
        "fund_name", "fundname",
    ],
    "security_type": [
        "security_type", "securitytype", "security type", "sec_type",
    ],
}


def detect_columns(headers: list[str]) -> dict[str, int]:
    """Map logical column names to header indices."""
    mapping: dict[str, int] = {}
    headers_lower = [h.strip().lower() for h in headers]

    for logical, patterns in COLUMN_PATTERNS.items():
        for i, h in enumerate(headers_lower):
            if any(p in h for p in patterns):
                if logical not in mapping:
                    mapping[logical] = i
                break
    return mapping


# ═══════════════════════════════════════════════════════════════════════════════
# Asset class tagging
# ═══════════════════════════════════════════════════════════════════════════════

def tag_asset_class(raw: str) -> AssetClassTag:
    r = raw.lower()
    if any(k in r for k in ["australian equities", "aust equities", "domestic equities"]):
        return AssetClassTag.AEQ
    if "aeq" == r.strip():
        return AssetClassTag.AEQ
    if any(k in r for k in [
        "international equities", "intl equities", "emerging market",
        "low volatility equit",
    ]):
        return AssetClassTag.IEQ
    if r.strip() in ("ieq", "emt"):
        return AssetClassTag.IEQ
    return AssetClassTag.OTHER


# ═══════════════════════════════════════════════════════════════════════════════
# Cash / residual exclusion
# ═══════════════════════════════════════════════════════════════════════════════

def is_cash_residual(row: ParsedRow) -> bool:
    for field_val in (row.security_type, row.entity_name):
        for pat in ("cash bucket", "cb trading", "fx residual", "cash residual"):
            if icontains(field_val, pat):
                return True
    return False


# ═══════════════════════════════════════════════════════════════════════════════
# AEQ classification
# ═══════════════════════════════════════════════════════════════════════════════

def classify_aeq(entity: str, mandate: str) -> tuple[ActivityType, SourceType, bool]:
    e = entity.lower()
    m = mandate.lower()

    # ACTIVE + EXTERNAL
    for pat in ["cooper investor", "greencape", "vinva", "alphinity",
                "platypus", "yarra", "ifm", "paradice"]:
        if pat in e or pat in m:
            return ActivityType.ACTIVE, SourceType.EXTERNAL, True

    # ACTIVE + INTERNAL
    for pat in ["obi", "opportunistic"]:
        if pat in e or pat in m:
            return ActivityType.ACTIVE, SourceType.INTERNAL, True
    if "vader" in e or "vader" in m:
        return ActivityType.ACTIVE, SourceType.INTERNAL, True
    if "imp" in m:
        return ActivityType.ACTIVE, SourceType.INTERNAL, True

    # PASSIVE + EXTERNAL
    for pat in ["state street", "ssga", "ssg"]:
        if pat in e or pat in m:
            return ActivityType.PASSIVE, SourceType.EXTERNAL, True

    # PASSIVE + INTERNAL
    for pat in ["asx20", "asx 20", "plug"]:
        if pat in e or pat in m:
            return ActivityType.PASSIVE, SourceType.INTERNAL, True

    return ActivityType.ACTIVE, SourceType.EXTERNAL, False  # unclassified


# ═══════════════════════════════════════════════════════════════════════════════
# IEQ classification
# ═══════════════════════════════════════════════════════════════════════════════

def classify_ieq(entity: str, mandate: str) -> tuple[ActivityType, SourceType, bool]:
    e = entity.lower()
    m = mandate.lower()

    # ACTIVE + EXTERNAL
    for pat in ["arrowstreet", "wellington", "sanders", "c worldwide",
                "c worldw", "orbis", "artisan", "wasatch", "rwc",
                "jennison", "gsam", "goldman sachs", "goldman"]:
        if pat in e or pat in m:
            return ActivityType.ACTIVE, SourceType.EXTERNAL, True

    # ACTIVE + INTERNAL
    for pat in ["elvis", "quality plus", "qualityplus", "qlpl"]:
        if pat in e or pat in m:
            return ActivityType.ACTIVE, SourceType.INTERNAL, True
    for pat in ["nvidia", "nvda"]:
        if pat in e or pat in m:
            return ActivityType.ACTIVE, SourceType.INTERNAL, True
    if "plug" in e or "plug" in m:
        return ActivityType.ACTIVE, SourceType.INTERNAL, True

    # PASSIVE + EXTERNAL — SSgA Low Carbon mandates
    for pat in ["state street", "ssga", "ssg", "low carbon",
                "lchosg", "lcwasg", "optimised"]:
        if pat in e or pat in m:
            return ActivityType.PASSIVE, SourceType.EXTERNAL, True

    # PASSIVE + INTERNAL — TRS / Swaps
    for pat in ["total return swap", "trs", "swap", "minvol", "msvl",
                "low vol", "lowvol", "minimum volatility"]:
        if pat in e or pat in m:
            return ActivityType.PASSIVE, SourceType.INTERNAL, True

    # IFM anomaly
    if "ifm" in e or "ifm" in m:
        print(f"  [WARNING] IFM Investors found in IEQ context — possible data anomaly.",
              file=sys.stderr)
        return ActivityType.ACTIVE, SourceType.EXTERNAL, True

    return ActivityType.ACTIVE, SourceType.EXTERNAL, False  # unclassified


# ═══════════════════════════════════════════════════════════════════════════════
# Output formatting
# ═══════════════════════════════════════════════════════════════════════════════

def print_separator(width: int = 90):
    print("-" * width)


def print_asset_class_table(label: str, date_str: str, entries: list[ManagerEntry]):
    entries.sort(key=lambda e: e.exposure_aud, reverse=True)

    total = sum(e.exposure_aud for e in entries)
    total_m = total / 1e6

    print()
    print_separator()
    print(f"  {label} as at {date_str}")
    print_separator()

    # Header
    print(f"{'Manager Name':<32} | {'Exposure ($m)':>14} | {'Active/Passive':<14} | {'Int/Ext':<14}")
    print_separator()

    unclassified = []
    for e in entries:
        exp_m = e.exposure_aud / 1e6
        tag = " [?]" if not e.classified else ""
        print(f"{(e.name + tag):<32} | {fmt_money(exp_m):>14} | {e.act.value:<14} | {e.src.value:<14}")
        if not e.classified:
            unclassified.append(e)

    print_separator()
    print(f"{'TOTAL ' + label:<32} | {fmt_money(total_m):>14} |{'':32}")
    print_separator()

    # Summaries
    active_sum = sum(e.exposure_aud for e in entries if e.act == ActivityType.ACTIVE)
    passive_sum = sum(e.exposure_aud for e in entries if e.act == ActivityType.PASSIVE)
    internal_sum = sum(e.exposure_aud for e in entries if e.src == SourceType.INTERNAL)
    external_sum = sum(e.exposure_aud for e in entries if e.src == SourceType.EXTERNAL)

    def pct(v):
        return (v / total * 100) if total > 0 else 0.0

    print(f"\n  SUMMARY — {label}:")
    print(f"    Active:    {fmt_money(active_sum / 1e6)}  ({fmt_pct(pct(active_sum))})")
    print(f"    Passive:   {fmt_money(passive_sum / 1e6)}  ({fmt_pct(pct(passive_sum))})")
    print(f"    Internal:  {fmt_money(internal_sum / 1e6)}  ({fmt_pct(pct(internal_sum))})")
    print(f"    External:  {fmt_money(external_sum / 1e6)}  ({fmt_pct(pct(external_sum))})")

    # 2x2 matrix
    ai = sum(e.exposure_aud for e in entries if e.act == ActivityType.ACTIVE and e.src == SourceType.INTERNAL)
    ae = sum(e.exposure_aud for e in entries if e.act == ActivityType.ACTIVE and e.src == SourceType.EXTERNAL)
    pi = sum(e.exposure_aud for e in entries if e.act == ActivityType.PASSIVE and e.src == SourceType.INTERNAL)
    pe = sum(e.exposure_aud for e in entries if e.act == ActivityType.PASSIVE and e.src == SourceType.EXTERNAL)

    def cell(v):
        return f"{fmt_money(v / 1e6)} ({fmt_pct(pct(v))})"

    print(f"\n  2x2 Matrix — {label}:")
    print(f"  {'':18}  | {'Internal':>18} | {'External':>18} | {'Total':>18}")
    print(f"  {'-' * 78}")
    print(f"  {'Active':<18}  | {cell(ai):>18} | {cell(ae):>18} | {cell(ai + ae):>18}")
    print(f"  {'Passive':<18}  | {cell(pi):>18} | {cell(pe):>18} | {cell(pi + pe):>18}")
    print(f"  {'Total':<18}  | {cell(ai + pi):>18} | {cell(ae + pe):>18} | {fmt_money(total_m) + ' (100.0%)':>18}")

    # Unclassified
    if unclassified:
        print(f"\n  [!] UNCLASSIFIED MANAGERS ({label}):")
        for u in unclassified:
            print(f"      - {u.name}  (mandate: {u.mandate_code}, exposure: {fmt_money(u.exposure_aud / 1e6)})")


# ═══════════════════════════════════════════════════════════════════════════════
# Validation
# ═══════════════════════════════════════════════════════════════════════════════

def run_validation(label: str, entries: list[ManagerEntry],
                   expected_total_low_bn: float, expected_total_high_bn: float,
                   expected_passive_low_pct: float, expected_passive_high_pct: float):
    total = sum(e.exposure_aud for e in entries)
    active_sum = sum(e.exposure_aud for e in entries if e.act == ActivityType.ACTIVE)
    passive_sum = sum(e.exposure_aud for e in entries if e.act == ActivityType.PASSIVE)
    internal_sum = sum(e.exposure_aud for e in entries if e.src == SourceType.INTERNAL)
    external_sum = sum(e.exposure_aud for e in entries if e.src == SourceType.EXTERNAL)
    has_unclassified = any(not e.classified for e in entries)

    total_bn = total / 1e9
    passive_pct = (passive_sum / total * 100) if total > 0 else 0.0

    # Manager concentration
    mgr_exp: dict[str, float] = defaultdict(float)
    for e in entries:
        mgr_exp[e.name] += e.exposure_aud

    print(f"\n  VALIDATION — {label}:")

    def check(ok: bool, msg: str):
        tag = "[PASS]" if ok else "[WARN]"
        print(f"    {tag} {msg}")

    check(expected_passive_low_pct <= passive_pct <= expected_passive_high_pct,
          f"{label} passive = {fmt_pct(passive_pct)} "
          f"(expected {fmt_pct(expected_passive_low_pct)}-{fmt_pct(expected_passive_high_pct)})")

    check(expected_total_low_bn <= total_bn <= expected_total_high_bn,
          f"Total {label} = ${total_bn:.1f}bn "
          f"(expected ${expected_total_low_bn:.0f}-${expected_total_high_bn:.0f}bn)")

    conc_ok = True
    for name, exp in mgr_exp.items():
        pct = (exp / total * 100) if total > 0 else 0.0
        if pct > 25.0:
            check(False, f"{name} = {fmt_pct(pct)} (exceeds 25% threshold)")
            conc_ok = False
    if conc_ok:
        check(True, f"No single manager exceeds 25% of {label}")

    check(not has_unclassified, "All rows classified (no unclassified managers)")

    ie_diff = abs((internal_sum + external_sum) - total)
    ap_diff = abs((active_sum + passive_sum) - total)
    check(ie_diff < 1.0, f"Sum of Internal + External = Total (diff: ${ie_diff / 1e6:.4f}m)")
    check(ap_diff < 1.0, f"Sum of Active + Passive = Total (diff: ${ap_diff / 1e6:.4f}m)")


# ═══════════════════════════════════════════════════════════════════════════════
# CSV export
# ═══════════════════════════════════════════════════════════════════════════════

def export_csv(filepath: str, aeq: list[ManagerEntry], ieq: list[ManagerEntry], date_str: str):
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "AssetClass", "ManagerName", "MandateCode", "ExposureAUD",
            "ExposureMillions", "ActivePassive", "InternalExternal", "Classified", "Date",
        ])
        for ac_label, entries in [("AEQ", aeq), ("IEQ", ieq)]:
            for e in entries:
                writer.writerow([
                    ac_label, e.name, e.mandate_code,
                    f"{e.exposure_aud:.2f}", f"{e.exposure_aud / 1e6:.1f}",
                    e.act.value, e.src.value,
                    "Y" if e.classified else "N", date_str,
                ])
    print(f"\n[CSV] Results exported to: {filepath}")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="VFMC Risk Manager — Active/Passive & Internal/External Split Calculator"
    )
    parser.add_argument("xlsx_path", help="Path to the VFMC RM export .xlsx file")
    parser.add_argument("date", nargs="?", default=None,
                        help="Date label (YYYYMMDD). Auto-detected from filename if omitted.")
    parser.add_argument("--csv", dest="csv_path", default=None,
                        help="Export results to a CSV file")
    args = parser.parse_args()

    xlsx_path = args.xlsx_path
    date_str = args.date
    csv_path = args.csv_path

    if not os.path.isfile(xlsx_path):
        print(f"[ERROR] File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    # Auto-detect date from filename
    if date_str is None:
        match = re.search(r"(\d{8})", Path(xlsx_path).name)
        date_str = match.group(1) if match else "UNKNOWN"

    print(f"[INFO] Reading: {xlsx_path}")
    print(f"[INFO] Date:    {date_str}")
    print()

    # ── Read Excel ───────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as ex:
        print(f"[ERROR] Failed to open xlsx file: {ex}", file=sys.stderr)
        sys.exit(1)

    ws = wb.active

    # ── Detect columns ───────────────────────────────────────────────────
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    col_map = detect_columns(headers)

    for logical in ("entity", "asset_class", "exposure"):
        if logical not in col_map and not (logical == "entity" and "mandate" in col_map):
            print(f"[ERROR] Cannot find required column: {logical}", file=sys.stderr)
            print(f"        Headers found: {headers}", file=sys.stderr)
            sys.exit(1)

    print("[INFO] Column mapping:")
    for logical, idx in col_map.items():
        print(f"  {logical:<16}: col {idx} ({headers[idx]})")
    print()

    # ── Parse rows ───────────────────────────────────────────────────────
    parsed: list[tuple[AssetClassTag, ParsedRow]] = []
    skipped_other = skipped_zero = skipped_cash = 0
    total_rows = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1

        def safe_str(logical: str) -> str:
            idx = col_map.get(logical)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        def safe_float(logical: str) -> float:
            idx = col_map.get(logical)
            if idx is None or idx >= len(row) or row[idx] is None:
                return 0.0
            val = row[idx]
            if isinstance(val, (int, float)):
                return float(val)
            try:
                return float(str(val).replace(",", "").strip())
            except (ValueError, TypeError):
                return 0.0

        pr = ParsedRow(
            entity_name=safe_str("entity"),
            asset_class_raw=safe_str("asset_class"),
            exposure=safe_float("exposure"),
            mandate_code=safe_str("mandate"),
            portfolio=safe_str("portfolio"),
            security_type=safe_str("security_type"),
        )

        ac = tag_asset_class(pr.asset_class_raw)
        if ac == AssetClassTag.OTHER:
            skipped_other += 1
            continue

        if abs(pr.exposure) < 0.01:
            skipped_zero += 1
            continue

        if is_cash_residual(pr):
            skipped_cash += 1
            continue

        parsed.append((ac, pr))

    wb.close()

    print(f"[INFO] Rows read:          {total_rows}")
    print(f"[INFO] Skipped (other AC): {skipped_other}")
    print(f"[INFO] Skipped (zero exp): {skipped_zero}")
    print(f"[INFO] Skipped (cash/FX):  {skipped_cash}")
    print(f"[INFO] Relevant rows:      {len(parsed)}")

    # ── Aggregate by manager ─────────────────────────────────────────────
    agg: dict[tuple[AssetClassTag, str, str], ManagerEntry] = {}

    for ac, pr in parsed:
        name = pr.entity_name if pr.entity_name else pr.mandate_code
        if not name:
            name = "UNKNOWN"
        key = (ac, name, pr.mandate_code)

        if key in agg:
            agg[key].exposure_aud += pr.exposure
        else:
            me = ManagerEntry(name=name, mandate_code=pr.mandate_code, ac=ac,
                              exposure_aud=pr.exposure)
            agg[key] = me

    # ── Classify ─────────────────────────────────────────────────────────
    aeq_entries: list[ManagerEntry] = []
    ieq_entries: list[ManagerEntry] = []
    unclassified_count = 0

    for (ac, name, mandate), me in agg.items():
        if ac == AssetClassTag.AEQ:
            act, src, classified = classify_aeq(me.name, me.mandate_code)
        else:
            act, src, classified = classify_ieq(me.name, me.mandate_code)

        me.act = act
        me.src = src
        me.classified = classified

        if not classified:
            unclassified_count += 1
            print(f"  [WARNING] Unclassified: \"{me.name}\" "
                  f"(mandate: {me.mandate_code}, exposure: {fmt_money(me.exposure_aud / 1e6)}, "
                  f"AC: {'AEQ' if ac == AssetClassTag.AEQ else 'IEQ'}) "
                  f"-> defaulting to Active+External", file=sys.stderr)

        if ac == AssetClassTag.AEQ:
            aeq_entries.append(me)
        else:
            ieq_entries.append(me)

    # ── Output ───────────────────────────────────────────────────────────
    if aeq_entries:
        print_asset_class_table("AEQ (Australian Equities)", date_str, aeq_entries)
        run_validation("AEQ", aeq_entries, 12.0, 14.0, 20.0, 35.0)
    else:
        print("\n[INFO] No AEQ data found in file.")

    if ieq_entries:
        print_asset_class_table("IEQ (International Equities)", date_str, ieq_entries)
        run_validation("IEQ", ieq_entries, 24.0, 28.0, 22.0, 38.0)
    else:
        print("\n[INFO] No IEQ data found in file.")

    # ── CSV export ───────────────────────────────────────────────────────
    if csv_path:
        export_csv(csv_path, aeq_entries, ieq_entries, date_str)

    # ── Final summary ────────────────────────────────────────────────────
    total_mgrs = len(aeq_entries) + len(ieq_entries)
    print()
    print_separator()
    print(f"  DONE. Managers classified: {total_mgrs - unclassified_count} / {total_mgrs}"
          f" | Unclassified: {unclassified_count}")
    print_separator()


if __name__ == "__main__":
    main()
