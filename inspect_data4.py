import openpyxl
import sys

wb = openpyxl.load_workbook('VFMC.Daily.20260430.CLIENT_DTF-Act3.AssetClass.xlsx', data_only=True)
ws = wb['RiskReport']

# Known column positions (0-indexed from col map)
COL = {
    'level': 2,     # Col B (1-indexed = 2)
    'name': 3,      # Col C
    'pv': 4,        # Col D (RM PV AUD / Total)
    'ac': 12,       # Col L
    'sec_type': 13,  # Col M
    'strategy': 136, # Col EJ
    'mgr_code': 141, # Col EO
    'mgr_name': 142, # Col EP
    'int_ext': 143,  # Col EQ
    'portfolio': 139, # Col EI
    'is_cash': 160,  # Col FD
    'sec_name': 146, # Col EP
}

max_row = ws.max_row
print(f"Total rows: {max_row}")

# Find all Level 0 rows (asset class level)
print("\n=== All Level 0 rows (Asset Classes) ===")
level0_rows = []
for row_num in range(22, max_row + 1):
    level = ws.cell(row=row_num, column=COL['level']).value
    if level == 0:
        name = ws.cell(row=row_num, column=COL['name']).value
        pv = ws.cell(row=row_num, column=COL['pv']).value
        print(f"  Row {row_num}: {name}  PV={pv}")
        level0_rows.append((row_num, str(name) if name else ''))
    if row_num % 5000 == 0:
        print(f"  ... scanning row {row_num}...", file=sys.stderr)

# Find AEQ and IEQ start rows
aeq_start = None
ieq_starts = []
for row_num, name in level0_rows:
    nl = name.lower()
    if 'australian equit' in nl:
        aeq_start = row_num
    if any(k in nl for k in ['international equit', 'emerging market', 'low volatility equit']):
        ieq_starts.append((row_num, name))

# Print Level 1 children for AEQ
if aeq_start:
    print(f"\n=== AEQ (Row {aeq_start}) — Level 1 children ===")
    for row_num in range(aeq_start + 1, min(aeq_start + 200, max_row + 1)):
        level = ws.cell(row=row_num, column=COL['level']).value
        if level is not None and level <= 0:
            break  # next asset class
        if level == 1:
            name = ws.cell(row=row_num, column=COL['name']).value
            pv = ws.cell(row=row_num, column=COL['pv']).value
            mgr = ws.cell(row=row_num, column=COL['mgr_name']).value
            mc = ws.cell(row=row_num, column=COL['mgr_code']).value
            ie = ws.cell(row=row_num, column=COL['int_ext']).value
            strat = ws.cell(row=row_num, column=COL['strategy']).value
            print(f"  Lv1 Row {row_num}: Name={name} PV={pv} Mgr={mgr} MC={mc} IE={ie} Strat={strat}")

# Print Level 1 children for first IEQ
if ieq_starts:
    row_num_start, ieq_name = ieq_starts[0]
    print(f"\n=== IEQ: {ieq_name} (Row {row_num_start}) — Level 1 children ===")
    for row_num in range(row_num_start + 1, min(row_num_start + 200, max_row + 1)):
        level = ws.cell(row=row_num, column=COL['level']).value
        if level is not None and level <= 0:
            break
        if level == 1:
            name = ws.cell(row=row_num, column=COL['name']).value
            pv = ws.cell(row=row_num, column=COL['pv']).value
            mgr = ws.cell(row=row_num, column=COL['mgr_name']).value
            mc = ws.cell(row=row_num, column=COL['mgr_code']).value
            ie = ws.cell(row=row_num, column=COL['int_ext']).value
            strat = ws.cell(row=row_num, column=COL['strategy']).value
            print(f"  Lv1 Row {row_num}: Name={name} PV={pv} Mgr={mgr} MC={mc} IE={ie} Strat={strat}")

wb.close()
