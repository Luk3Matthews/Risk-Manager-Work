import openpyxl

wb = openpyxl.load_workbook('VFMC.Daily.20260430.CLIENT_DTF-Act3.AssetClass.xlsx', read_only=True, data_only=True)

ws = wb['RiskReport']

# Print all non-empty content in rows 15-25
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
    non_empty = [(j, str(c).strip()) for j, c in enumerate(row) if c is not None]
    if non_empty:
        print(f"Row {i} ({len(row)} cols): {non_empty[:15]}...")

# Now print full row 20 content
print("\n=== Full Row 20 ===")
for i, row in enumerate(ws.iter_rows(min_row=20, max_row=20, values_only=True), start=20):
    all_vals = [(j, str(c).strip()) for j, c in enumerate(row) if c is not None]
    for idx, val in all_vals:
        print(f"  Col {idx}: {val}")
    print(f"  Total non-empty cells: {len(all_vals)}")

# Print rows 21-25
print("\n=== Rows 21-25 ===")
for i, row in enumerate(ws.iter_rows(min_row=21, max_row=25, values_only=True), start=21):
    non_empty = [(j, str(c).strip()) for j, c in enumerate(row) if c is not None]
    if non_empty:
        print(f"Row {i}: {non_empty[:15]}...")

# Also check what's in the data area (rows 26-35)
print("\n=== Rows 26-35 ===")
for i, row in enumerate(ws.iter_rows(min_row=26, max_row=35, values_only=True), start=26):
    non_empty = [(j, str(c).strip()) for j, c in enumerate(row) if c is not None]
    if non_empty:
        print(f"Row {i}: {non_empty[:15]}...")

wb.close()
