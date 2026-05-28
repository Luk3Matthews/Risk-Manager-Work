import openpyxl

# Use read_only=False to handle merged cells properly
wb = openpyxl.load_workbook('VFMC.Daily.20260430.CLIENT_DTF-Act3.AssetClass.xlsx', data_only=True)

# Get column headers from the metadata sheet (row 8)
meta_ws = wb['RiskReport_metaData(statistic']
meta_headers = [str(cell.value).strip() if cell.value else '' for cell in meta_ws[8]]
print("=== Headers from metadata sheet (row 8) ===")
for i, h in enumerate(meta_headers):
    if h and any(k in h.lower() for k in ['manager', 'asset class', 'pv aud', 'int_ext',
                                            'iscash', 'sectype', 'portfolio', 'level',
                                            'attribute', 'strategy', 'instrument']):
        print(f"  Col {i}: {h}")

# Now scan RiskReport for actual data
ws = wb['RiskReport']
print(f"\n=== RiskReport: scanning rows 11-35 ===")
for row_num in range(11, 36):
    row_cells = [cell.value for cell in ws[row_num]]
    non_empty = [(j, v) for j, v in enumerate(row_cells) if v is not None]
    if non_empty:
        print(f"Row {row_num} ({len(non_empty)} non-empty): {non_empty[:10]}")

# Also try reading specific cells at known column positions
# From metadata: Col 1=Level, 2=Attribute, 3=RM PV AUD, 11=Asset Class, 60=ManagerName
print(f"\n=== RiskReport: key columns from rows 11-50 ===")
count = 0
for row_num in range(11, 51):
    level = ws.cell(row=row_num, column=2).value  # Col B = idx 1
    attr = ws.cell(row=row_num, column=3).value    # Col C = idx 2
    pv = ws.cell(row=row_num, column=4).value      # Col D = idx 3
    ac = ws.cell(row=row_num, column=12).value     # Col L = idx 11
    mgr = ws.cell(row=row_num, column=61).value    # Col BI = idx 60
    mc = ws.cell(row=row_num, column=60).value     # Col BH = idx 59
    ie = ws.cell(row=row_num, column=62).value     # Col BJ = idx 61
    port = ws.cell(row=row_num, column=58).value   # Col BF = idx 57

    if pv is not None or level is not None:
        print(f"  Row {row_num}: Lv={level} Attr={attr} PV={pv} AC={ac} Mgr={mgr} MC={mc} IE={ie} Port={port}")
        count += 1
        if count >= 15:
            break

wb.close()
